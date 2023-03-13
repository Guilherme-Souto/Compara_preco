import re
import schedule
import time
from pathlib import Path
from email_sender.emailsender import Send_email
from webscraping.webscrap import create_browser, pick_price, pick_product_title, BeautifulSoup
from openpyxl import load_workbook

# caminho da tabela que ele pega os dados
TABLE_PATH = Path(__file__).parent / 'Produtos.xlsx'

# Tudo fica dentro dessa func para ser executado pelo schedule.


def main():

    # classe que herda do send_email e sobrescreve os metodos
    class Novo_Send_email(Send_email):
        def __init__(self):
            super().__init__()
            self.subject = f'Desconto Amazon {title_subject}'

        def email_text(self):
            text_email = f'{title} está custando {final_price}'
            return text_email

        # aqui tem que colocar o email que vc quer ser notificado
        def collect_email_receiver(self):
            self.email_receiver = 'CHANGE ME'
            self.email_receiver.strip
            self._verify_email_received()

        def run(self):
            self.collect_email_receiver()
            self.smtp_config()
            self.email_text()
            self.create_mime()
            self.sending_email()

    links = []
    prices = []

    # func que tem as configs do excel
    def excel():
        workbook = load_workbook(TABLE_PATH)

        sheet_name = 'Produtos'

        worksheet = workbook[sheet_name]
        return worksheet

    table = excel()

    # for que percorre a tabela
    for a, b in table.iter_rows(min_row=2):
        if a.value != None:
            links.append(a.value)
        if b.value != None:
            prices.append(b.value)

    # for que faz a pesquisa na internet
    for index, value in enumerate(links):
        create_browser()
        browser = create_browser()
        browser.get(value)

        soup = BeautifulSoup(browser.page_source, "html.parser")
        title = pick_product_title(soup)
        web_price = pick_price(soup)

        final_price = ''.join([n for n in web_price if n.isdigit()])
        final_price = round(float(final_price)/100, 2)

        semi_title = title.strip()[:10]
        title_subject = re.sub('[^A-Za-z0-9]+', '', semi_title)
        print(f'Analisando {title_subject}')
        desired_price = prices[index]

        if final_price <= desired_price:
            c1 = Novo_Send_email()
            c1.run()


# a func que executa o codigo a cada hora
schedule.every(1).hours.do(main)

while 1:
    schedule.run_pending()
    time.sleep(1800)
